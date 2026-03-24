import serial
import csv
import os
import time
import shutil
import ctypes
import threading
import math
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
import pystray
from PIL import Image, ImageDraw

# --- 1. CONFIGURATION ---
COM_PORT = 'COM4'  # Note: Users should update this to match their specific receiver port
BAUD_RATE = 9600 

# GitHub Version: Dynamic pathing for zero-configuration deployment
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Primary local save location (creates an 'Exported_Data' folder next to the script)
PRIMARY_FOLDER = os.path.join(BASE_DIR, "Exported_Data")

# Secondary Backup Location (creates a backup folder next to the script)
BACKUP_FOLDER = os.path.join(BASE_DIR, "Exported_Data_Backup")

# The Master Usage Log File
USAGE_LOG_FILE = os.path.join(PRIMARY_FOLDER, "Usage_History_Log.csv")

# Ensure directories exist on startup
os.makedirs(PRIMARY_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)

CONFIG_FILE = os.path.join(BASE_DIR, "logger_directory.csv") 

# Production Mode Timeout (5 minutes)
TIMEOUT_SECONDS = 5 * 60

# --- 2. MEMORY & GRAPHING DATA ---
last_seen_times = {}
active_sessions = {}
plot_data = {} 
time_data = {} 
unsaved_data_buffer = {}  

# --- 3. LOAD THE DIRECTORY ---
def load_transmitters():
    transmitters = {}
    if not os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Address", "Logger_Name"]) 
            writer.writerow(["00001", "Thermolog-01_ZoneA"]) 
            writer.writerow(["00002", "Thermolog-02_ZoneB"]) 
            writer.writerow(["00003", "Thermolog-03_Incubator"])
        print(f"Created a new dictionary file: {CONFIG_FILE}.")
        
    with open(CONFIG_FILE, mode='r') as file:
        reader = csv.reader(file)
        next(reader, None)  
        for row in reader:
            if len(row) >= 2:
                addr = str(row[0]).strip().zfill(5) 
                name = str(row[1]).strip()
                transmitters[addr] = name
    return transmitters

# --- 4. EXCEL POST-PROCESSING & DUAL LOGGING ---
def create_excel_report(csv_path):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Run Data"
        
        min_temp = float('inf')
        max_temp = float('-inf')
        
        with open(csv_path, 'r') as file:
            reader = csv.reader(file)
            header = next(reader)
            ws.append(header)
            
            for row in reader:
                if len(row) < 5: 
                    continue
                
                # 1. Format Date
                date_val = row[0]
                
                # 2. Format Data Point
                try: dp_val = int(row[1])
                except ValueError: dp_val = row[1]
                    
                # 3. Format Time (SECRETLY COMBINED FOR MATH, SHOWN AS TIME)
                try: 
                    # Recombine Date and Time so Excel's ScatterChart doesn't crash
                    time_val = datetime.strptime(f"{date_val} {row[2]}", "%Y-%m-%d %H:%M:%S")
                except ValueError: 
                    time_val = row[2]
                    
                # 4. Format Process Temp & Track Min/Max
                try: 
                    proc_temp = float(row[3])
                    if proc_temp < min_temp: min_temp = proc_temp
                    if proc_temp > max_temp: max_temp = proc_temp
                except ValueError: 
                    proc_temp = row[3]
                    
                # 5. Format Ambient Temp
                try: amb_temp = float(row[4])
                except ValueError: amb_temp = row[4]
                    
                ws.append([date_val, dp_val, time_val, proc_temp, amb_temp])
        
        # Explicitly format the Time column (Column C) to ONLY show hh:mm:ss visually
        for row in range(2, ws.max_row + 1):
            ws[f"C{row}"].number_format = "hh:mm:ss"
            
        chart = ScatterChart()
        chart.title = "Process Temperature Profile"
        chart.style = 13 
        chart.y_axis.title = "Temperature (°C)"
        chart.x_axis.title = "Time"
        chart.scatterStyle = "line"
        
        chart.legend = None 
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.number_format = 'hh:mm:ss'
        chart.y_axis.number_format = '0'
        
        # Apply +/- 20% logic and snap bounds to the nearest multiple of 5 
        if min_temp != float('inf') and max_temp != float('-inf'):
            y_min = min_temp - (abs(min_temp) * 0.20)
            y_max = max_temp + (abs(max_temp) * 0.20)
            if y_min == y_max:
                y_min -= 5
                y_max += 5
            
            # Using float() ensures openpyxl perfectly reads the bounds
            chart.y_axis.scaling.min = float(math.floor(y_min / 5.0) * 5)
            chart.y_axis.scaling.max = float(math.ceil(y_max / 5.0) * 5)
        
        x_data = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
        y_data = Reference(ws, min_col=4, min_row=2, max_row=ws.max_row)
        
        series = Series(values=y_data, xvalues=x_data, title="Process Temp (°C)")
        chart.series.append(series)
        
        chart.width = 24
        chart.height = 12
        ws.add_chart(chart, "G2") 
        
        excel_path = csv_path.replace('.csv', '.xlsx')
        wb.save(excel_path)
        
        if os.path.exists(excel_path):
            try: os.remove(csv_path)
            except PermissionError: pass
            return excel_path
            
    except PermissionError:
        return None
    except Exception:
        return None

def log_data(name, data_point_num, process_temp, ambient_temp, session_timestamp):
    log_date = datetime.now().strftime("%Y-%m-%d")
    log_time = datetime.now().strftime("%H:%M:%S")
    
    filename = f"{name}_Run_{session_timestamp}.csv"
    
    month_object = datetime.strptime(session_timestamp, "%Y-%m-%d_%I-%M%p")
    month_folder = month_object.strftime("%Y-%b")
    
    primary_logger_folder = os.path.join(PRIMARY_FOLDER, name, month_folder)
    os.makedirs(primary_logger_folder, exist_ok=True)
    primary_path = os.path.join(primary_logger_folder, filename)
    
    if primary_path not in unsaved_data_buffer:
        unsaved_data_buffer[primary_path] = []
        
    unsaved_data_buffer[primary_path].append([log_date, data_point_num, log_time, process_temp, ambient_temp])
    
    primary_exists = os.path.isfile(primary_path)
    
    try:
        with open(primary_path, mode='a', newline='') as file:
            writer = csv.writer(file)
            if not primary_exists:
                writer.writerow(["Date", "Data Point", "Time", "Process Temp (C)", "Ambient Temp (C)"])
            
            for row in unsaved_data_buffer[primary_path]:
                writer.writerow(row)
                
            unsaved_data_buffer[primary_path] = []
            
    except PermissionError:
        pass 
    except Exception:
        pass 

# --- 5. SYSTEM TRAY CONTROLS ---
def create_tray_icon():
    image = Image.new('RGB', (64, 64), color='#2b3035')
    dc = ImageDraw.Draw(image)
    dc.rectangle((16, 16, 48, 48), fill='#39ff14')
    return image

def quit_app(icon, item):
    icon.stop()
    os._exit(0)  

def run_tray():
    icon = pystray.Icon("MissionControl", create_tray_icon(), "Mission Control", menu=pystray.Menu(
        pystray.MenuItem("Shutdown Mission Control", quit_app)
    ))
    icon.run()

# --- 6. MAIN PROGRAM & DASHBOARD ---
def main():
    try: ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)
    except Exception: pass 

    threading.Thread(target=run_tray, daemon=True).start()

    active_loggers = load_transmitters()
    
    last_csv_mtime = os.path.getmtime(CONFIG_FILE) if os.path.exists(CONFIG_FILE) else time.time()
    last_config_check = time.time()
    
    theme_colors = ['#39ff14', '#00ffff', '#ff00ff', '#ffff00', '#ff9900', '#0088ff', '#ff3333', '#b026ff', '#00ff99', '#ffffff']
    assigned_colors = {}
    color_idx = 0
    
    plt.ion() 
    plt.style.use('dark_background')
    fig = plt.figure(figsize=(10, 8), facecolor='#2b3035') 
    fig.canvas.manager.set_window_title('Mission Control - Temperature Monitor')

    try:
        def disable_close(): pass 
        fig.canvas.manager.window.protocol("WM_DELETE_WINDOW", disable_close)
    except Exception: pass 

    try:
        ser = serial.Serial(COM_PORT, BAUD_RATE, timeout=0.1) 

        while True:
            current_time = time.time()
            data_received = False
            
            if (current_time - last_config_check) > 10: 
                try:
                    current_mtime = os.path.getmtime(CONFIG_FILE)
                    if current_mtime != last_csv_mtime:
                        active_loggers = load_transmitters() 
                        last_csv_mtime = current_mtime
                except Exception:
                    pass 
                last_config_check = current_time
            
            if ser.in_waiting > 0:
                if ser.read() == b'\x7E':
                    packet = ser.read(15) 
                    
                    if len(packet) == 15:
                        addr_msb = packet[3]
                        addr_lsb = packet[4]
                        address = f"{addr_msb:02X}{addr_lsb:02X}".zfill(5) 
                        
                        temp_msb = packet[8]
                        temp_lsb = packet[9]
                        raw_process_f = (temp_msb << 8) | temp_lsb
                        
                        amb_msb = packet[10]
                        amb_lsb = packet[11]
                        raw_ambient_f = ((amb_msb << 8) | amb_lsb) / 10.0
                        
                        if raw_process_f > 5000: continue 
                            
                        final_process_temp = (raw_process_f - 32) * (5.0 / 9.0)
                        final_ambient_temp = (raw_ambient_f - 32) * (5.0 / 9.0)
                        
                        if address in active_loggers:
                            device_name = active_loggers[address]
                            data_received = True
                            
                            if address not in last_seen_times or (current_time - last_seen_times[address]) > TIMEOUT_SECONDS:
                                session_id = datetime.now().strftime("%Y-%m-%d_%I-%M%p")
                                active_sessions[address] = session_id
                                plot_data[address] = [] 
                                time_data[address] = []
                                
                                if address not in assigned_colors:
                                    assigned_colors[address] = theme_colors[color_idx % len(theme_colors)]
                                    color_idx += 1
                                
                                try:
                                    log_exists = os.path.isfile(USAGE_LOG_FILE)
                                    with open(USAGE_LOG_FILE, mode='a', newline='') as uf:
                                        u_writer = csv.writer(uf)
                                        if not log_exists:
                                            u_writer.writerow(["Date", "Time", "Logger_Used", "Run_ID"])
                                        
                                        run_date = datetime.now().strftime("%Y-%m-%d")
                                        run_time = datetime.now().strftime("%I:%M:%S %p")
                                        u_writer.writerow([run_date, run_time, device_name, session_id])
                                except Exception:
                                    pass
                            
                            last_seen_times[address] = current_time
                            current_dp_count = len(plot_data[address]) + 1
                            
                            log_data(device_name, current_dp_count, final_process_temp, final_ambient_temp, active_sessions[address])
                            plot_data[address].append(final_process_temp)
                            time_data[address].append(datetime.now())
            
            timeouts_to_remove = []
            for addr in list(active_sessions.keys()):
                if (current_time - last_seen_times[addr]) > TIMEOUT_SECONDS:
                    device_name = active_loggers[addr]
                    session_id = active_sessions[addr]
                    
                    month_object = datetime.strptime(session_id, "%Y-%m-%d_%I-%M%p")
                    month_folder = month_object.strftime("%Y-%b")
                    
                    primary_path = os.path.join(PRIMARY_FOLDER, device_name, month_folder, f"{device_name}_Run_{session_id}.csv")
                    if os.path.exists(primary_path):
                        final_excel_path = create_excel_report(primary_path)
                        
                        if final_excel_path and os.path.exists(final_excel_path):
                            try:
                                backup_dir = os.path.join(BACKUP_FOLDER, device_name, month_folder)
                                os.makedirs(backup_dir, exist_ok=True)
                                shutil.copy2(final_excel_path, backup_dir)
                            except Exception:
                                pass
                                
                    timeouts_to_remove.append(addr)
            
            for addr in timeouts_to_remove:
                del active_sessions[addr]
                if addr in plot_data: del plot_data[addr]
                if addr in time_data: del time_data[addr]

            if data_received or len(timeouts_to_remove) > 0:
                fig.clf() 
                active_count = len(active_sessions)
                
                if active_count == 0:
                    fig.text(0.5, 0.5, "Waiting for loggers to transmit...", ha='center', va='center', fontsize=16, color='#39ff14')
                else:
                    for i, (addr, session_id) in enumerate(active_sessions.items()):
                        name = active_loggers[addr]
                        ax = fig.add_subplot(active_count, 1, i + 1)
                        ax.set_facecolor('#2b3035') 
                        
                        y_data = plot_data[addr]
                        t_data = time_data[addr]
                        line_color = assigned_colors[addr]
                        
                        ax.plot(t_data, y_data, label=f"Current: {y_data[-1]:.2f}°C", color=line_color, linewidth=2.5)
                        
                        locator = mdates.AutoDateLocator()
                        formatter = mdates.ConciseDateFormatter(locator)
                        ax.xaxis.set_major_locator(locator)
                        ax.xaxis.set_major_formatter(formatter)
                        
                        current_min = min(y_data)
                        current_max = max(y_data)
                        y_min = current_min - (abs(current_min) * 0.20)
                        y_max = current_max + (abs(current_max) * 0.20)
                        
                        if y_min == y_max:
                            y_min -= 5
                            y_max += 5
                        ax.set_ylim(y_min, y_max)
                        
                        ax2 = ax.twiny()
                        ax2.set_xlim(1, max(2, len(y_data)))
                        ax2.xaxis.set_major_locator(MaxNLocator(integer=True))
                        
                        ax.set_xlabel("Time", color='white', fontweight='bold', labelpad=8)
                        ax2.set_xlabel("Data Point", color='white', fontweight='bold', labelpad=8)
                        
                        ax2.tick_params(colors='white')
                        
                        ax.set_title(f"{name} (Run ID: {session_id})", fontsize=12, fontweight='bold', color='white', pad=35) 
                        ax.set_ylabel("Temp (°C)", color='white', fontweight='bold')
                        ax.tick_params(colors='white')
                        ax.grid(True, linestyle='--', alpha=0.3, color='gray')
                        
                        legend = ax.legend(loc="upper left", facecolor='#2b3035', edgecolor='gray')
                        for text in legend.get_texts(): text.set_color("white")
                
                plt.tight_layout()
            
            plt.pause(0.05) 
                        
    except serial.SerialException as e:
        # Save error log dynamically in the app directory instead of the Desktop
        error_log_path = os.path.join(BASE_DIR, "MissionControl_ErrorLog.txt")
        with open(error_log_path, "w") as f:
            f.write(f"Connection error: {e}\nMake sure TC Central is closed, AND that Mission Control isn't already running!")
    except KeyboardInterrupt: pass

if __name__ == '__main__':
    main()
